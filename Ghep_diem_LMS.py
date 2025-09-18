import os
import re
import pandas as pd
from openpyxl import load_workbook
import customtkinter as ctk
from tkinter import filedialog, messagebox

# ======================= HỖ TRỢ LẤY HEADER / CỘT ==========================

def read_excel_with_header_detect(file_path):
    """Đọc Excel, tự động tìm dòng chứa 'Mã SV' để làm header (thử nhiều dòng đầu)."""
    for i in range(0, 15):
        try:
            df = pd.read_excel(file_path, header=i)
        except:
            continue
        cols = [str(c).lower() for c in df.columns]
        if any("mã sv" in c for c in cols):
            return df
    # fallback
    return pd.read_excel(file_path)

def find_column(df, keywords):
    """Tìm cột chứa 1 trong các keyword (không phân biệt hoa thường)."""
    for c in df.columns:
        c_low = str(c).lower()
        for kw in keywords:
            if kw in c_low:
                return c
    return None

def find_tbc_dtp_column(df):
    """Tìm chính xác cột TBC ĐTP (*) (ưu tiên cột chứa cả 'tbc' và 'đtp')."""
    for c in df.columns:
        c_low = str(c).lower().replace('\n', ' ')
        if 'tbc' in c_low and 'đtp' in c_low:
            return c
    # một số file có dấu khác, thử tìm 'đtp' hoặc 'tbc' chứa 'đtp'
    for c in df.columns:
        c_low = str(c).lower().replace('\n', ' ')
        if 'đtp' in c_low or 'tbc đtp' in c_low:
            return c
    return None

# ======================= HỖ TRỢ LẤY MÃ MÔN / NHÓM ==========================

def extract_subject_group_from_cell(file_path):
    """
    Đọc ô C5/C6 lấy text, xử lý một vài dạng:
    - "Tiếng Anh 1 (LCE315) - 06" -> subject=LCE315, group=06
    - Có thể gặp "(251-LCE315-01)" -> subject=LCE315, group=01
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        # thử C5 trước, nếu rỗng thử C6
        raw = ws['C5'].value
        if raw is None or str(raw).strip() == "":
            raw = ws['C6'].value
        if raw is None:
            return None, None

        text = str(raw).strip()

        # 1) nếu có phần trong ngoặc dạng 251-LCE315-01 hoặc LCE315
        paren = re.search(r'\((.*?)\)', text)
        subject_code = None
        group_code = None
        if paren:
            inside = paren.group(1).strip()
            # nếu inside có dấu '-' phân tách như 251-LCE315-01
            parts = inside.split('-')
            if len(parts) >= 3:
                # parts e.g. ['251','LCE315','01']
                subject_code = parts[1].strip()
                group_code = parts[2].strip()
            elif len(parts) == 2:
                # e.g. 'LCE315-01' hoặc '251-LCE315'
                # nếu phần đầu là số, lấy phần sau làm subject (251-LCE315)
                if parts[0].isdigit():
                    subject_code = parts[1].strip()
                else:
                    subject_code = parts[0].strip()
                    group_code = parts[1].strip()
            else:
                # chỉ có 1 phần trong ngoặc, có thể là LCE315
                subject_code = inside.strip()

        # 2) nếu không có ngoặc, thử tách pattern " - 06" ở cuối
        if group_code is None:
            gm = re.search(r'-\s*([A-Za-z0-9]+)\s*$', text)
            if gm:
                group_code = gm.group(1).strip()

        # 3) nếu subject_code vẫn None, tìm mã dạng LEX123 trong chuỗi
        if subject_code is None:
            sm = re.search(r'\b([A-Za-z]{2,}\d{2,})\b', text)
            if sm:
                subject_code = sm.group(1).strip()

        return subject_code, group_code
    except:
        return None, None

def extract_subject_group_from_filename(filename):
    """
    Tách mã môn + nhóm từ tên file nếu không tìm được trong file.
    Ví dụ: "...(251-LCE315-01).xlsx" => LCE315, 01
    """
    base = os.path.basename(filename)
    m = re.search(r'\((?:\d+-)?([A-Za-z0-9]+)-(\d+)\)', base, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2)
    # thử tìm pattern LCE315-01 không trong ngoặc
    m2 = re.search(r'([A-Za-z]{2,}\d{2,})[-_ ]+(\d{1,3})', base)
    if m2:
        return m2.group(1), m2.group(2)
    return None, None

def extract_subject_group(file_path):
    """Tổng hợp: thử cell trước, nếu ko có thử filename."""
    sub, grp = extract_subject_group_from_cell(file_path)
    if (sub is None or sub == "") or (grp is None or grp == ""):
        # thử filename
        fn_sub, fn_grp = extract_subject_group_from_filename(file_path)
        if sub is None or sub == "":
            sub = fn_sub
        if grp is None or grp == "":
            grp = fn_grp
    # chuẩn hóa None -> ""
    if sub is None:
        sub = ""
    if grp is None:
        grp = ""
    return sub, grp

# ======================= KIỂM TRA MÃ SV HỢP LỆ ==========================

def is_probably_masv(s):
    """Trả True nếu chuỗi có vẻ là Mã SV, loại trừ dòng tiêu đề/ghi chú."""
    if pd.isna(s):
        return False
    st = str(s).strip()
    if st == "":
        return False
    # loại trừ các dòng ghi chú
    if re.search(r'(số sv|môn học|cbgd|nhóm|tbc|điều kiện)', st, re.IGNORECASE):
        return False
    # nếu chỉ chứa chữ số hoặc chữ+số hoặc có dấu gạch, coi là mã SV hợp lệ
    if re.match(r'^[A-Za-z0-9\-/]+$', st):
        # tránh những chuỗi quá ngắn như '01' (thường là nhãn nhóm)
        if len(st) <= 2 and st.isdigit():
            return False
        return True
    return False

# ======================= GHÉP FILE ==========================

def merge_files():
    folder_in = folder_in_var.get()
    folder_out = folder_out_var.get()
    file_out = file_out_var.get().strip()

    if not folder_in or not os.path.isdir(folder_in):
        messagebox.showerror("Lỗi", "Bạn chưa chọn thư mục chứa file Excel!")
        return
    if not folder_out or not os.path.isdir(folder_out):
        messagebox.showerror("Lỗi", "Bạn chưa chọn thư mục lưu file kết quả!")
        return
    if file_out == "":
        file_out = "Tong_Hop_Diem.xlsx"
    if not file_out.lower().endswith('.xlsx'):
        file_out += '.xlsx'

    excel_files = [f for f in os.listdir(folder_in) if f.lower().endswith(('.xlsx', '.xls'))]
    if not excel_files:
        messagebox.showwarning("Không có file", "Thư mục không có file Excel!")
        return

    all_data = []
    for f in excel_files:
        path = os.path.join(folder_in, f)

        # lấy subject & group (cell trước, filename sau)
        subject_code, group_code = extract_subject_group(path)

        # đọc bảng chính
        try:
            df = read_excel_with_header_detect(path)
        except:
            continue

        # tìm cột Mã SV và cột TBC ĐTP
        col_ma_sv = find_column(df, ['mã sv', 'masv'])
        col_tbc = find_tbc_dtp_column(df)

        if col_ma_sv is None or col_tbc is None:
            # không đủ thông tin để lấy dữ liệu -> bỏ qua file
            continue

        try:
            temp = df[[col_ma_sv, col_tbc]].copy()
        except:
            continue

        # lọc: chỉ dòng có khả năng là SV thật
        temp = temp[temp[col_ma_sv].apply(is_probably_masv)]
        # thêm cột thông tin môn & nhóm, đảm bảo có kiểu str
        temp = temp.copy()
        temp.rename(columns={col_ma_sv: 'Mã SV', col_tbc: 'Điểm TBC'}, inplace=True)
        temp['Mã môn học'] = str(subject_code) if subject_code is not None else ""
        temp['Nhóm'] = str(group_code) if group_code is not None else ""
        # đưa cột theo thứ tự mong muốn
        cols_order = ['Mã SV', 'Mã môn học', 'Nhóm', 'Điểm TBC']
        # nếu có các cột khác thì giữ nguyên sau đó
        temp = temp[[c for c in cols_order if c in temp.columns] + [c for c in temp.columns if c not in cols_order]]

        all_data.append(temp)

    if all_data:
        result = pd.concat(all_data, ignore_index=True)
        # đảm bảo thứ tự final
        final_cols = ['Mã SV', 'Mã môn học', 'Nhóm', 'Điểm TBC']
        final_cols = [c for c in final_cols if c in result.columns] + [c for c in result.columns if c not in final_cols]
        result = result[final_cols]
        out_path = os.path.join(folder_out, file_out)
        result.to_excel(out_path, index=False)
        messagebox.showinfo("Thành công", f"Đã lưu file kết quả tại:\n{out_path}")
    else:
        messagebox.showwarning("Không có dữ liệu", "Không ghép được dữ liệu hợp lệ từ các file Excel!")

# ======================= GIAO DIỆN ==========================

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

root = ctk.CTk()
root.title("Ghép điểm - TNT - LMS")
root.geometry("750x270")

folder_in_var = ctk.StringVar()
folder_out_var = ctk.StringVar()
file_out_var = ctk.StringVar(value="Tong_Hop_Diem.xlsx")

def select_folder_in():
    fol = filedialog.askdirectory(title="Chọn thư mục chứa file Excel")
    if fol:
        folder_in_var.set(fol)

def select_folder_out():
    fol = filedialog.askdirectory(title="Chọn thư mục lưu file kết quả")
    if fol:
        folder_out_var.set(fol)

frame = ctk.CTkFrame(root)
frame.pack(padx=20, pady=20, fill="both", expand=True)

ctk.CTkLabel(frame, text="Thư mục chứa file Excel:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
ctk.CTkEntry(frame, textvariable=folder_in_var, width=430).grid(row=0, column=1, padx=5, pady=5)
ctk.CTkButton(frame, text="Chọn...", command=select_folder_in).grid(row=0, column=2, padx=5, pady=5)

ctk.CTkLabel(frame, text="Thư mục lưu file kết quả:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
ctk.CTkEntry(frame, textvariable=folder_out_var, width=430).grid(row=1, column=1, padx=5, pady=5)
ctk.CTkButton(frame, text="Chọn...", command=select_folder_out).grid(row=1, column=2, padx=5, pady=5)

ctk.CTkLabel(frame, text="Tên file kết quả (.xlsx):").grid(row=2, column=0, sticky="w", padx=5, pady=5)
ctk.CTkEntry(frame, textvariable=file_out_var, width=430).grid(row=2, column=1, padx=5, pady=5)

ctk.CTkButton(frame, text="Ghép dữ liệu", command=merge_files, fg_color="green").grid(row=3, column=1, pady=20)

root.mainloop()
