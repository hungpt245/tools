import os
import re
import pandas as pd
from openpyxl import load_workbook
import customtkinter as ctk
from tkinter import filedialog, messagebox

# ======================= HÀM HỖ TRỢ ==========================

def read_excel_with_header_detect(file_path):
    """Đọc Excel, tự động tìm dòng có Mã SV hoặc TBC để làm header"""
    for i in range(0, 10):  # thử 10 dòng đầu tiên
        try:
            df = pd.read_excel(file_path, header=i)
        except:
            continue
        cols = [str(c).lower() for c in df.columns]
        if any("mã sv" in c for c in cols) or any("tbc" in c for c in cols):
            return df
    # Nếu không tìm thấy header, trả về mặc định
    return pd.read_excel(file_path)

def find_column(df, keywords):
    """Tìm tên cột chứa 1 trong các keyword (không phân biệt hoa thường)"""
    for c in df.columns:
        c_low = str(c).lower()
        for kw in keywords:
            if kw in c_low:
                return c
    return None

def extract_subject_group_from_cell(file_path):
    """Đọc ô C5 (hoặc C6) trong file Excel và tách mã môn học + nhóm"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        # thử ô C5 trước, nếu rỗng thử ô C6
        cell_value = ws['C5'].value or ws['C6'].value
        if not cell_value:
            return None, None

        text = str(cell_value)
        # Lấy mã môn học trong dấu ngoặc
        subject_match = re.search(r'\((.*?)\)', text)
        subject_code = subject_match.group(1).strip() if subject_match else None

        # Lấy nhóm sau dấu -
        group_match = re.search(r'-\s*(\w+)', text)
        group_code = group_match.group(1).strip() if group_match else None

        return subject_code, group_code
    except:
        return None, None

def is_valid_masv(value):
    """Kiểm tra xem value có phải là mã sinh viên hợp lệ không"""
    if pd.isna(value):
        return False
    s = str(value).strip()
    # loại bỏ các dòng có chữ 'điều kiện'
    if 'điều kiện' in s.lower():
        return False
    # mã SV hợp lệ: chữ số/chữ cái/ký tự gạch (-,/)
    if re.match(r'^[A-Za-z0-9\-/]+$', s):
        return True
    return False

# ======================= CHỨC NĂNG GHÉP FILE ==========================

def merge_files():
    folder_in = folder_in_var.get()
    folder_out = folder_out_var.get()
    output_name = file_name_var.get().strip()

    if not folder_in or not os.path.isdir(folder_in):
        messagebox.showerror("Lỗi", "Bạn chưa chọn thư mục chứa file Excel!")
        return
    if not folder_out or not os.path.isdir(folder_out):
        messagebox.showerror("Lỗi", "Bạn chưa chọn thư mục lưu file kết quả!")
        return
    if not output_name:
        messagebox.showerror("Lỗi", "Bạn chưa nhập tên file kết quả!")
        return

    excel_files = [f for f in os.listdir(folder_in) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        messagebox.showwarning("Không có file", "Thư mục không có file Excel!")
        return

    all_data = []
    for f in excel_files:
        file_path = os.path.join(folder_in, f)

        # lấy mã môn học + nhóm từ ô C5/C6
        subject_code, group_code = extract_subject_group_from_cell(file_path)

        try:
            df = read_excel_with_header_detect(file_path)
        except:
            continue

        col_ma_sv = find_column(df, ['mã sv', 'masv'])
        col_tbc = find_column(df, ['tbc', 'đtp'])

        if col_ma_sv is None or col_tbc is None:
            # bỏ qua file không đủ cột
            continue

        try:
            temp = df[[col_ma_sv, col_tbc]].copy()
        except:
            continue

        # lọc bỏ các dòng mà cột Mã SV bị trống
        temp = temp[temp[col_ma_sv].notna()]

        # lọc bỏ các dòng thống kê không phải sinh viên
        temp = temp[temp[col_ma_sv].apply(is_valid_masv)]

        temp.rename(columns={col_ma_sv: 'Mã SV', col_tbc: 'Điểm TBC'}, inplace=True)
        temp['Mã môn học'] = subject_code
        temp['Nhóm'] = group_code

        all_data.append(temp)

    if all_data:
        result = pd.concat(all_data, ignore_index=True)

        # Đảm bảo tên file có đuôi .xlsx
        if not output_name.lower().endswith('.xlsx'):
            output_name += '.xlsx'

        save_path = os.path.join(folder_out, output_name)
        result.to_excel(save_path, index=False)
        messagebox.showinfo("Thành công", f"Đã lưu file kết quả tại:\n{save_path}")
    else:
        messagebox.showwarning("Không có dữ liệu", "Không ghép được dữ liệu từ các file Excel!")

# ======================= GIAO DIỆN CUSTOMTKINTER ==========================

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

root = ctk.CTk()
root.title("Ghép điểm - TNT - AQ")
root.geometry("750x270")

folder_in_var = ctk.StringVar()
folder_out_var = ctk.StringVar()
file_name_var = ctk.StringVar(value="Tong_Hop_Diem.xlsx")  # tên mặc định

def select_folder_in():
    folder_selected = filedialog.askdirectory(title="Chọn thư mục chứa file Excel")
    if folder_selected:
        folder_in_var.set(folder_selected)

def select_folder_out():
    folder_selected = filedialog.askdirectory(title="Chọn thư mục lưu file kết quả")
    if folder_selected:
        folder_out_var.set(folder_selected)

frame = ctk.CTkFrame(root)
frame.pack(padx=20, pady=20, fill="both", expand=True)

lbl_in = ctk.CTkLabel(frame, text="Thư mục chứa file Excel:")
lbl_in.grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_in = ctk.CTkEntry(frame, textvariable=folder_in_var, width=400)
entry_in.grid(row=0, column=1, padx=5, pady=5)
btn_in = ctk.CTkButton(frame, text="Chọn...", command=select_folder_in)
btn_in.grid(row=0, column=2, padx=5, pady=5)

lbl_out = ctk.CTkLabel(frame, text="Thư mục lưu file kết quả:")
lbl_out.grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_out = ctk.CTkEntry(frame, textvariable=folder_out_var, width=400)
entry_out.grid(row=1, column=1, padx=5, pady=5)
btn_out = ctk.CTkButton(frame, text="Chọn...", command=select_folder_out)
btn_out.grid(row=1, column=2, padx=5, pady=5)

lbl_file = ctk.CTkLabel(frame, text="Tên file kết quả (.xlsx):")
lbl_file.grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_file = ctk.CTkEntry(frame, textvariable=file_name_var, width=400)
entry_file.grid(row=2, column=1, padx=5, pady=5)

btn_merge = ctk.CTkButton(frame, text="Ghép dữ liệu", command=merge_files, fg_color="green")
btn_merge.grid(row=3, column=1, padx=5, pady=30)

root.mainloop()
